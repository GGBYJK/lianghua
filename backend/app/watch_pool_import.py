from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "监控品种",
    "监控周期",
    "头部到左颈，头部到右颈最小高度",
    "左颈到左肩，右颈到右肩最小价差",
    "启用关键区间趋势评分",
    "阻挡区间",
    "支撑区间",
    "交易时间段",
    "监控开关",
]
ALLOWED_TIMEFRAMES = {"1m", "3m", "5m", "15m", "30m", "60m", "1h"}
DEFAULT_TRADING_SESSIONS = "day,night"


@dataclass
class ImportIssue:
    row: int
    reason: str
    symbol: str | None = None
    timeframe: str | None = None
    field: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "row": self.row,
            "symbol": self.symbol,
            "timeframe": self.timeframe,
            "field": self.field,
            "reason": self.reason,
        }


def parse_watch_pool_excel(content: bytes, contract_lookup: dict[str, str]) -> tuple[list[dict[str, Any]], list[ImportIssue]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Excel 文件无法读取：{exc}") from exc

    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空")

    headers = [normalize_text(value) for value in rows[0][: len(EXPECTED_HEADERS)]]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Excel 表头必须为：{', '.join(EXPECTED_HEADERS)}")

    items: list[dict[str, Any]] = []
    errors: list[ImportIssue] = []
    seen: set[tuple[str, str]] = set()
    normalized_contracts = {normalize_symbol(symbol): name for symbol, name in contract_lookup.items()}

    for index, row in enumerate(rows[1:], start=2):
        values = list(row[: len(EXPECTED_HEADERS)])
        if all(is_blank(value) for value in values):
            continue

        raw = dict(zip(EXPECTED_HEADERS, values))
        symbol = normalize_symbol(raw.get("监控品种"))
        timeframe = normalize_text(raw.get("监控周期"))
        row_errors: list[ImportIssue] = []

        if not symbol:
            row_errors.append(ImportIssue(index, "监控品种不能为空", None, timeframe or None, "监控品种"))
        elif symbol not in normalized_contracts:
            row_errors.append(ImportIssue(index, "监控品种必须存在于合约中心", symbol, timeframe or None, "监控品种"))
        if timeframe not in ALLOWED_TIMEFRAMES:
            row_errors.append(ImportIssue(index, "监控周期必须是 1m、3m、5m、15m、30m、60m 或 1h", symbol or None, timeframe or None, "监控周期"))

        monitor_minutes = timeframe_to_monitor_minutes(timeframe) if timeframe in ALLOWED_TIMEFRAMES else 1

        min_height = parse_float(raw.get("头部到左颈，头部到右颈最小高度"), default=0.0)
        if min_height is None or min_height < 0:
            row_errors.append(ImportIssue(index, "头部到左颈，头部到右颈最小高度必须是大于等于 0 的数字", symbol or None, timeframe or None, "头部到左颈，头部到右颈最小高度"))

        min_shoulder_height = parse_float(raw.get("左颈到左肩，右颈到右肩最小价差"), default=0.0)
        if min_shoulder_height is None or min_shoulder_height < 0:
            row_errors.append(ImportIssue(index, "左颈到左肩，右颈到右肩最小价差必须是大于等于 0 的数字", symbol or None, timeframe or None, "左颈到左肩，右颈到右肩最小价差"))

        key_zone_enabled = parse_bool(raw.get("启用关键区间趋势评分"), default=False)
        if key_zone_enabled is None:
            row_errors.append(ImportIssue(index, "启用关键区间趋势评分必须是 true/false、是/否、开启/关闭 或 1/0", symbol or None, timeframe or None, "启用关键区间趋势评分"))

        resistance_zone = parse_zone(raw.get("阻挡区间"))
        if resistance_zone is None:
            row_errors.append(ImportIssue(index, "阻挡区间必须是数字区间，例如 3500-3520", symbol or None, timeframe or None, "阻挡区间"))

        support_zone = parse_zone(raw.get("支撑区间"))
        if support_zone is None:
            row_errors.append(ImportIssue(index, "支撑区间必须是数字区间，例如 3300-3320", symbol or None, timeframe or None, "支撑区间"))

        trading_sessions = normalize_trading_sessions(raw.get("交易时间段"))
        if not trading_sessions:
            row_errors.append(ImportIssue(index, "交易时间段必须是 day、night 或 day,night", symbol or None, timeframe or None, "交易时间段"))

        enabled = parse_bool(raw.get("监控开关"), default=True)
        if enabled is None:
            row_errors.append(ImportIssue(index, "监控开关必须是 true/false、是/否、开启/关闭 或 1/0", symbol or None, timeframe or None, "监控开关"))

        key = (symbol, timeframe)
        if symbol and timeframe and timeframe in ALLOWED_TIMEFRAMES:
            if key in seen:
                row_errors.append(ImportIssue(index, "Excel 内部重复，已保留第一条", symbol, timeframe, None))
            seen.add(key)

        if row_errors:
            errors.extend(row_errors)
            continue

        items.append(
            {
                "name": normalized_contracts.get(symbol) or symbol,
                "symbol": symbol,
                "timeframe": timeframe,
                "enabled": bool(enabled),
                "monitor_minutes": int(monitor_minutes),
                "trading_sessions": trading_sessions or DEFAULT_TRADING_SESSIONS,
                "min_head_to_neck_height": float(min_height),
                "min_shoulder_to_neck_height": float(min_shoulder_height),
                "enable_key_zone_trend_score": bool(key_zone_enabled),
                "resistance_zone_min": float(resistance_zone[0]) if resistance_zone else 0.0,
                "resistance_zone_max": float(resistance_zone[1]) if resistance_zone else 0.0,
                "support_zone_min": float(support_zone[0]) if support_zone else 0.0,
                "support_zone_max": float(support_zone[1]) if support_zone else 0.0,
                "_row": index,
            }
        )

    return items, errors


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_symbol(value: Any) -> str:
    return normalize_text(value)


def is_blank(value: Any) -> bool:
    return normalize_text(value) == ""


def parse_int(value: Any, default: int | None = None) -> int | None:
    if is_blank(value):
        return default
    try:
        number = float(str(value).strip())
    except ValueError:
        return None
    if not number.is_integer():
        return None
    return int(number)


def parse_float(value: Any, default: float | None = None) -> float | None:
    if is_blank(value):
        return default
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def parse_bool(value: Any, default: bool = True) -> bool | None:
    if is_blank(value):
        return default
    text = normalize_text(value).lower()
    if text in {"true", "1", "是", "开启", "开", "yes", "y"}:
        return True
    if text in {"false", "0", "否", "关闭", "关", "no", "n"}:
        return False
    return None


def parse_zone(value: Any) -> tuple[float, float] | None:
    if is_blank(value):
        return (0.0, 0.0)
    parts: list[float] = []
    for part in normalize_text(value).replace("，", ",").replace("、", ",").replace("~", ",").replace("-", ",").split(","):
        text = part.strip()
        if not text:
            continue
        try:
            number = float(text)
        except ValueError:
            return None
        if number <= 0:
            return None
        parts.append(number)
    if not parts:
        return (0.0, 0.0)
    if len(parts) == 1:
        return (parts[0], parts[0])
    return (min(parts[0], parts[1]), max(parts[0], parts[1]))


def timeframe_to_monitor_minutes(timeframe: str) -> int:
    if timeframe.endswith("m"):
        return max(1, int(timeframe[:-1] or 1))
    if timeframe.endswith("h"):
        return max(1, int(timeframe[:-1] or 1) * 60)
    return 1


def normalize_trading_sessions(value: Any) -> str | None:
    if is_blank(value):
        return DEFAULT_TRADING_SESSIONS
    text = normalize_text(value).replace("，", ",").replace("、", ",")
    aliases = {
        "白天": "day",
        "日盘": "day",
        "day": "day",
        "夜间": "night",
        "夜盘": "night",
        "night": "night",
    }
    sessions: list[str] = []
    for part in [item.strip() for item in text.split(",") if item.strip()]:
        key = aliases.get(part.lower()) or aliases.get(part)
        if key is None:
            return None
        if key not in sessions:
            sessions.append(key)
    return ",".join(sessions) if sessions else None
