from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .market_client import contract_to_variety


@dataclass(frozen=True)
class ProductCostImportIssue:
    row: int
    reason: str


def _header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("不是有效数字") from exc


def _parse_fee_part(value: str) -> tuple[str, Decimal] | None:
    rate = re.search(r"万分之\s*(\d+(?:\.\d+)?)", value)
    if rate:
        return "TURNOVER_RATE", _decimal(rate.group(1)) / Decimal("10000")
    fixed = re.search(r"(\d+(?:\.\d+)?)\s*元\s*/?\s*手", value)
    if fixed:
        return "PER_LOT", _decimal(fixed.group(1))
    return None


def parse_fee_description(value: Any) -> dict[str, Decimal | str | None]:
    description = str(value or "").strip()
    if not description:
        raise ValueError("手续费为空")
    parts = [part.strip() for part in re.split(r"[，,；;]", description) if part.strip()]
    ordinary: tuple[str, Decimal] | None = None
    close_today: tuple[str, Decimal] | None = None
    for part in parts:
        parsed = _parse_fee_part(part)
        if parsed is None:
            continue
        if "平今" in part or "日内" in part:
            close_today = parsed
        elif ordinary is None:
            ordinary = parsed
    if ordinary is None:
        raise ValueError(f"无法解析手续费：{description}")
    return {
        "fee_mode": ordinary[0],
        "fee_value": ordinary[1],
        "fee_close_today_mode": close_today[0] if close_today else None,
        "fee_close_today_value": close_today[1] if close_today else None,
        "fee_description": description,
    }


def parse_product_cost_excel(content: bytes) -> tuple[list[dict[str, Any]], list[ProductCostImportIssue]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return [], [ProductCostImportIssue(1, "Excel为空")]
    columns = {_header(value): index for index, value in enumerate(header_row)}
    required = {"品种", "品种英文", "手续费", "保证金比率"}
    missing = sorted(required - set(columns))
    if missing:
        return [], [ProductCostImportIssue(1, f"缺少列：{'、'.join(missing)}")]

    items: list[dict[str, Any]] = []
    issues: list[ProductCostImportIssue] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        try:
            raw_symbol = str(row[columns["品种英文"]] or "").strip()
            symbol = contract_to_variety(raw_symbol)
            if symbol is None:
                raise ValueError("品种英文不是有效期货品种代码")
            margin_rate = _decimal(row[columns["保证金比率"]])
            if margin_rate > 1 and margin_rate <= 100:
                margin_rate /= Decimal("100")
            if margin_rate <= 0 or margin_rate > 1:
                raise ValueError("保证金比率必须在 0 到 1 之间")
            item = {
                "symbol": symbol.lower(),
                "exchange": symbol.split(".", 1)[0],
                "name": str(row[columns["品种"]] or "").strip(),
                "margin_rate": margin_rate,
                **parse_fee_description(row[columns["手续费"]]),
            }
            items.append(item)
        except (IndexError, ValueError) as exc:
            issues.append(ProductCostImportIssue(row_number, str(exc)))
    return items, issues


def parse_contract_specs_excel(content: bytes) -> tuple[list[dict[str, Any]], list[ProductCostImportIssue]]:
    """Parse the full contract-parameter workbook used by the admin screen."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return [], [ProductCostImportIssue(1, "Excel为空")]
    columns = {_header(value): index for index, value in enumerate(header_row)}
    required = {"品种", "名称", "合约乘数", "最小变动", "保证金率", "手续费"}
    missing = sorted(required - set(columns))
    if missing:
        return [], [ProductCostImportIssue(1, f"缺少列：{'、'.join(missing)}")]

    items: list[dict[str, Any]] = []
    issues: list[ProductCostImportIssue] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        try:
            raw_symbol = str(row[columns["品种"]] or "").strip()
            symbol = contract_to_variety(raw_symbol)
            if symbol is None:
                raise ValueError("品种不是有效期货品种代码")
            multiplier = _decimal(row[columns["合约乘数"]])
            price_tick = _decimal(row[columns["最小变动"]])
            margin_rate = _decimal(row[columns["保证金率"]])
            if multiplier <= 0:
                raise ValueError("合约乘数必须大于 0")
            if price_tick <= 0:
                raise ValueError("最小变动必须大于 0")
            if margin_rate > 1 and margin_rate <= 100:
                margin_rate /= Decimal("100")
            if margin_rate <= 0 or margin_rate > 1:
                raise ValueError("保证金率必须在 0 到 1 之间")
            fees = parse_fee_description(row[columns["手续费"]])
            items.append({
                "symbol": symbol.lower(),
                "exchange": symbol.split(".", 1)[0],
                "name": str(row[columns["名称"]] or "").strip(),
                "multiplier": multiplier,
                "price_tick": price_tick,
                "margin_rate": margin_rate,
                "fee_mode": fees["fee_mode"],
                "fee_value": fees["fee_value"],
                "fee_close_today_mode": fees["fee_close_today_mode"],
                "fee_close_today_value": fees["fee_close_today_value"],
                "enabled": True,
            })
        except (IndexError, ValueError) as exc:
            issues.append(ProductCostImportIssue(row_number, str(exc)))
    return items, issues
