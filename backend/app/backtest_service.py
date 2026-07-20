from __future__ import annotations

import asyncio
import json
import logging
import os
from dataclasses import replace
from datetime import datetime
from decimal import ROUND_FLOOR, Decimal
from io import BytesIO
from typing import Any
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook

from .backtest_store import (
    BacktestLeaseLostError,
    add_backtest_error,
    all_backtest_orders,
    claim_next_backtest_run,
    finish_backtest_run,
    get_backtest_run,
    is_backtest_cancel_requested,
    replace_backtest_summaries,
    save_backtest_orders,
    save_backtest_series,
    touch_backtest_heartbeat,
    update_backtest_progress,
)
from .config import load_head_shoulder_config
from .market_client import fetch_kline_from_market
from .strategy import add_ma_columns, add_macd_columns, find_pivots, prepare_chart_payload, scan_head_shoulders, signal_direction
from .market_client import contract_to_variety
from .trading_service import DEFAULT_SLIPPAGE_TICKS, _fee, _fill_price, _round_price, decimal_value
from .trading_store import get_contract_spec


logger = logging.getLogger("app.backtest")
BACKTEST_HEARTBEAT_SECONDS = float(os.getenv("BACKTEST_HEARTBEAT_SECONDS", "10"))


async def _heartbeat_backtest_run(
    run_id: str,
    worker_id: str,
    stop_event: asyncio.Event,
    lease_lost: asyncio.Event,
) -> None:
    while not stop_event.is_set():
        try:
            owned = await asyncio.to_thread(touch_backtest_heartbeat, run_id, worker_id)
        except Exception:
            logger.exception("backtest heartbeat failed: run=%s worker=%s", run_id, worker_id)
        else:
            if not owned:
                lease_lost.set()
                return
        try:
            await asyncio.wait_for(stop_event.wait(), timeout=BACKTEST_HEARTBEAT_SECONDS)
        except asyncio.TimeoutError:
            pass


def _backtest_shape_config(symbol: str, timeframe: str):
    return replace(
        load_head_shoulder_config(symbol, timeframe),
        min_head_to_neck_height=10.0,
        min_shoulder_to_neck_height=4.0,
        apply_ma60_pattern_penalty=True,
    )


def _analyze_market(
    frame: pd.DataFrame,
    hourly: pd.DataFrame,
    daily: pd.DataFrame,
    symbol: str,
    timeframe: str,
    entry_conditions: list[str],
    other_entry_conditions: list[str],
    min_pattern_score: int,
    min_trend_score: int,
    other_min_pattern_score: int,
    other_max_trend_score: int,
) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, Any]]:
    normalized = frame.copy().reset_index(drop=True)
    normalized["datetime"] = pd.to_datetime(normalized["datetime"])
    hourly = hourly.copy().reset_index(drop=True)
    daily = daily.copy().reset_index(drop=True)
    hourly["datetime"] = pd.to_datetime(hourly["datetime"])
    daily["datetime"] = pd.to_datetime(daily["datetime"])
    config = replace(
        _backtest_shape_config(symbol, timeframe),
        max_signal_age_bars=0,
        pullback_min_pattern_score=other_min_pattern_score,
        pullback_max_trend_score=other_max_trend_score,
    )
    signals = scan_head_shoulders(
        normalized,
        symbol,
        timeframe,
        config,
        hourly_df=hourly,
        daily_df=daily,
        include_right_neck_trigger=True,
    )
    selected = _filter_backtest_signals(
        [signal.to_dict() for signal in signals],
        entry_conditions,
        other_entry_conditions,
        min_pattern_score,
        min_trend_score,
        other_min_pattern_score,
        other_max_trend_score,
    )
    enriched = add_macd_columns(add_ma_columns(normalized, config), config)
    pivots = find_pivots(enriched, left=config.pivot_left, right=config.pivot_right)
    chart = prepare_chart_payload(enriched, pivots, signals, config, timeframe=timeframe)
    return normalized, selected, chart


def _score(signal: dict[str, Any], field: str) -> int:
    try:
        return int(signal.get(field) or 0)
    except (TypeError, ValueError):
        return 0


def _filter_backtest_signals(
    signals: list[dict[str, Any]],
    entry_conditions: list[str],
    other_entry_conditions: list[str],
    min_pattern_score: int,
    min_trend_score: int,
    other_min_pattern_score: int,
    other_max_trend_score: int,
) -> list[dict[str, Any]]:
    scored_conditions = set(entry_conditions)
    other_conditions = set(other_entry_conditions)
    selected: list[dict[str, Any]] = []
    for signal in signals:
        condition = f"{signal.get('pattern')}:{signal.get('alert_type')}"
        if condition in other_conditions and (
            _score(signal, "pattern_score") >= other_min_pattern_score
            and _score(signal, "score") <= other_max_trend_score
        ):
            selected.append(signal)
        elif condition in scored_conditions and (
            _score(signal, "pattern_score") >= min_pattern_score
            and _score(signal, "score") >= min_trend_score
        ):
            selected.append(signal)
    return selected


def _signal_time(signal: dict[str, Any]) -> str | None:
    return signal.get("retest_time") or signal.get("break_time") or signal.get("right_shoulder", {}).get("time")


def _signal_key(signal: dict[str, Any]) -> str:
    return "|".join([
        str(signal.get("symbol", "")),
        str(signal.get("timeframe", "")),
        str(signal.get("pattern", "")),
        str(signal.get("alert_type", "")),
        str(signal.get("head", {}).get("time", "")),
        str(_signal_time(signal) or ""),
    ])


def _contract_for_symbol(symbol: str) -> dict[str, Any] | None:
    candidates = [contract_to_variety(symbol), symbol, symbol.split(".")[-1]]
    for candidate in candidates:
        if not candidate:
            continue
        spec = get_contract_spec(candidate)
        if spec and spec.get("enabled"):
            return spec
    return None


def _position_key(symbol: str) -> str:
    return (contract_to_variety(symbol) or symbol).lower()


def _position_quantity(
    initial_capital: Decimal,
    single_symbol_position_pct: Decimal | None,
    entry_price: Decimal,
    contract: dict[str, Any] | None,
    position_sizing_mode: str = "PERCENT",
    single_symbol_lots: int | None = None,
) -> int:
    """Return fixed lots or the largest whole-lot position within the margin budget."""
    if contract is None or entry_price <= 0:
        return 0
    try:
        multiplier = decimal_value(contract["multiplier"])
        margin_rate = decimal_value(contract["margin_rate"])
    except (KeyError, TypeError, ValueError):
        return 0
    margin_per_lot = entry_price * multiplier * margin_rate
    if multiplier <= 0 or margin_rate <= 0 or margin_per_lot <= 0:
        return 0
    if position_sizing_mode == "FIXED_LOTS":
        return single_symbol_lots or 0
    if single_symbol_position_pct is None:
        return 0
    budget = initial_capital * single_symbol_position_pct / Decimal("100")
    return int((budget / margin_per_lot).to_integral_value(rounding=ROUND_FLOOR))


def _position_margin(order: dict[str, Any], contract: dict[str, Any] | None) -> Decimal:
    if contract is None or not order.get("entry_price") or not order.get("quantity"):
        return Decimal("0")
    try:
        return (
            decimal_value(order["entry_price"])
            * Decimal(str(order["quantity"]))
            * decimal_value(contract["multiplier"])
            * decimal_value(contract["margin_rate"])
        )
    except (KeyError, TypeError, ValueError):
        return Decimal("0")


def _trading_session_key(timestamp: pd.Timestamp) -> tuple[str, pd.Timestamp] | None:
    value = pd.Timestamp(timestamp)
    minute = value.hour * 60 + value.minute
    if minute >= 21 * 60:
        return "night", value.normalize()
    if minute < 9 * 60:
        return "night", (value - pd.Timedelta(days=1)).normalize()
    if 9 * 60 <= minute < 15 * 60:
        return "day", value.normalize()
    return None


def _session_penultimate_exit_index(timestamps: pd.Series, entry_index: int) -> int | None:
    session_key = _trading_session_key(pd.Timestamp(timestamps.iloc[entry_index]))
    if session_key is None:
        return None

    session_indices: list[int] = []
    for index in range(entry_index, len(timestamps)):
        if _trading_session_key(pd.Timestamp(timestamps.iloc[index])) != session_key:
            break
        session_indices.append(index)

    # A boundary must be present in the data; otherwise the order remains incomplete.
    if len(session_indices) < 2 or session_indices[-1] == len(timestamps) - 1:
        return None
    return session_indices[-2]


def _target_price(rule: dict[str, Any], signal: dict[str, Any], entry: float, stop: float, direction: str) -> float | None:
    metrics = signal.get("pattern_metrics") or {}
    if rule["type"] == "PATTERN_TARGET":
        value = metrics.get("target")
        return float(value) if value is not None else None
    if rule["type"] == "RR":
        distance = abs(entry - stop) * float(rule["multiplier"])
    else:
        qtr = signal.get("qtr")
        if qtr is None or float(qtr) <= 0:
            return None
        distance = float(qtr) * float(rule["multiplier"])
    return entry + distance if direction == "LONG" else entry - distance


def _stop_price(signal: dict[str, Any], direction: str, qtr_multiplier: float) -> float | None:
    try:
        qtr = float(signal.get("qtr") or 0)
        if qtr <= 0:
            return None
        is_pullback = str(signal.get("alert_type", "")).endswith("_pullback")
        point_key = "right_neck" if is_pullback else "right_shoulder"
        point = signal.get(point_key) or {}
        anchor = float(point["price"])
    except (KeyError, TypeError, ValueError):
        return None
    distance = qtr * qtr_multiplier
    return anchor - distance if direction == "LONG" else anchor + distance


def _simulate_order(
    *,
    run_id: str,
    series_id: str,
    frame: pd.DataFrame,
    signal: dict[str, Any],
    rule: dict[str, Any],
    max_holding_bars: int | None,
    contract: dict[str, Any] | None,
    entry_condition: str = "mixed",
    stop_loss_qtr_multiplier: float = 0.5,
    initial_capital: Decimal = Decimal("1000000"),
    single_symbol_position_pct: Decimal | None = Decimal("10"),
    position_sizing_mode: str = "PERCENT",
    single_symbol_lots: int | None = None,
    no_overnight: bool = False,
) -> dict[str, Any]:
    direction = signal_direction(str(signal["pattern"]), str(signal["alert_type"]))
    base = {
        "id": str(uuid4()),
        "run_id": run_id,
        "series_id": series_id,
        "rule_key": rule["key"],
        "rule_label": rule["label"],
        "signal_key": _signal_key(signal),
        "symbol": signal["symbol"],
        "timeframe": signal["timeframe"],
        "pattern": signal["pattern"],
        "alert_type": signal["alert_type"],
        "entry_condition": entry_condition,
        "direction": direction,
        "score": int(signal.get("pattern_score") or signal.get("score") or 0),
        "quantity": 0,
        "status": "INVALID",
        "exit_reason": None,
        "entry_time": None,
        "exit_time": None,
        "entry_price": None,
        "stop_price": None,
        "target_price": None,
        "exit_price": None,
        "gross_pnl": None,
        "net_pnl": None,
        "fees": None,
        "slippage": None,
        "r_multiple": None,
        "holding_bars": 0,
        "mfe_r": None,
        "mae_r": None,
        "cost_available": contract is not None,
        "signal_json": json.dumps(signal, ensure_ascii=False, separators=(",", ":")),
    }
    signal_time = _signal_time(signal)
    if not signal_time:
        return base
    timestamps = pd.to_datetime(frame["datetime"])
    matches = timestamps[timestamps == pd.Timestamp(signal_time)].index
    if len(matches) == 0:
        return base
    entry_index = int(matches[0])
    base["entry_time"] = timestamps.iloc[entry_index].to_pydatetime()
    # Signals are only actionable after this bar has closed.
    entry = float(frame.iloc[entry_index]["close"])
    stop = _stop_price(signal, direction, stop_loss_qtr_multiplier)
    if stop is None:
        return base
    target = _target_price(rule, signal, entry, stop, direction)
    if target is None:
        return base
    valid_prices = stop < entry < target if direction == "LONG" else target < entry < stop
    if not valid_prices:
        return base

    tick = decimal_value(contract["price_tick"]) if contract else Decimal("0")
    stop_decimal = _round_price(Decimal(str(stop)), tick) if contract else Decimal(str(stop))
    target_decimal = _round_price(Decimal(str(target)), tick) if contract else Decimal(str(target))
    open_side = "BUY" if direction == "LONG" else "SELL"
    close_side = "SELL" if direction == "LONG" else "BUY"
    entry_decimal = Decimal(str(entry))
    if contract:
        entry_fill, entry_slip = _fill_price(entry_decimal, open_side, tick, DEFAULT_SLIPPAGE_TICKS)
    else:
        entry_fill, entry_slip = entry_decimal, Decimal("0")
    quantity = _position_quantity(
        initial_capital,
        single_symbol_position_pct,
        entry_decimal,
        contract,
        position_sizing_mode,
        single_symbol_lots,
    )
    if contract and quantity < 1:
        return base
    base["quantity"] = quantity
    risk = abs(entry_fill - stop_decimal)
    if risk <= 0:
        return base

    session_exit_index = _session_penultimate_exit_index(timestamps, entry_index) if no_overnight else None
    if session_exit_index is not None and session_exit_index <= entry_index:
        return base

    available = len(frame) - entry_index - 1
    end_index = min(len(frame) - 1, entry_index + max_holding_bars) if max_holding_bars is not None else len(frame) - 1
    if session_exit_index is not None:
        end_index = min(end_index, session_exit_index)
    forced_by_session = session_exit_index is not None and end_index == session_exit_index
    exit_reason: str | None = None
    exit_index: int | None = None
    exit_quote: Decimal | None = None
    max_favorable = Decimal("0")
    max_adverse = Decimal("0")
    for index in range(entry_index + 1, end_index + 1):
        row = frame.iloc[index]
        low = Decimal(str(row["low"]))
        high = Decimal(str(row["high"]))
        if direction == "LONG":
            max_favorable = max(max_favorable, high - entry_fill)
            max_adverse = max(max_adverse, entry_fill - low)
            stop_hit = low <= stop_decimal
            target_hit = high >= target_decimal
        else:
            max_favorable = max(max_favorable, entry_fill - low)
            max_adverse = max(max_adverse, high - entry_fill)
            stop_hit = high >= stop_decimal
            target_hit = low <= target_decimal
        if stop_hit:
            exit_reason, exit_index, exit_quote = "STOP_LOSS", index, Decimal(str(row["close"]))
            break
        if target_hit:
            exit_reason, exit_index, exit_quote = "TAKE_PROFIT", index, Decimal(str(row["close"]))
            break

    if exit_index is None:
        if not forced_by_session and (max_holding_bars is None or available < max_holding_bars):
            base.update({
                "status": "INCOMPLETE",
                "entry_price": entry_fill,
                "stop_price": stop_decimal,
                "target_price": target_decimal,
                "holding_bars": max(0, available),
                "mfe_r": max_favorable / risk,
                "mae_r": max_adverse / risk,
            })
            return base
        exit_reason, exit_index = ("SESSION_EXIT" if forced_by_session else "TIME_EXIT"), end_index
        exit_quote = Decimal(str(frame.iloc[exit_index]["close"]))

    assert exit_quote is not None and exit_index is not None
    if contract:
        exit_fill, exit_slip = _fill_price(exit_quote, close_side, tick, DEFAULT_SLIPPAGE_TICKS)
        multiplier = decimal_value(contract["multiplier"])
        gross_per_lot = (exit_fill - entry_fill) * multiplier if direction == "LONG" else (entry_fill - exit_fill) * multiplier
        gross = gross_per_lot * quantity
        fees = _fee(entry_fill, quantity, contract, "OPEN") + _fee(exit_fill, quantity, contract, "CLOSE")
        net = gross - fees
        slippage = (entry_slip + exit_slip) * multiplier * quantity
    else:
        exit_fill, exit_slip = exit_quote, Decimal("0")
        gross = fees = net = slippage = None
    price_pnl = exit_fill - entry_fill if direction == "LONG" else entry_fill - exit_fill
    base.update({
        "status": "CLOSED",
        "exit_reason": exit_reason,
        "exit_time": timestamps.iloc[exit_index].to_pydatetime(),
        "entry_price": entry_fill,
        "stop_price": stop_decimal,
        "target_price": target_decimal,
        "exit_price": exit_fill,
        "gross_pnl": gross,
        "net_pnl": net,
        "fees": fees,
        "slippage": slippage,
        "r_multiple": price_pnl / risk,
        "holding_bars": exit_index - entry_index,
        "mfe_r": max_favorable / risk,
        "mae_r": max_adverse / risk,
    })
    return base


def _head_key(signal: dict[str, Any]) -> str:
    head = signal.get("head") or {}
    return "|".join([
        str(signal.get("pattern", "")),
        str(head.get("time", "")),
    ])


def _signal_timestamp(signal: dict[str, Any]) -> pd.Timestamp | None:
    value = _signal_time(signal)
    if not value:
        return None
    try:
        return pd.Timestamp(value)
    except (TypeError, ValueError):
        return None


def _has_higher_score(signal: dict[str, Any], previous_signal: dict[str, Any]) -> bool:
    return (
        _score(signal, "pattern_score") > _score(previous_signal, "pattern_score")
        or _score(signal, "score") > _score(previous_signal, "score")
    )


def _simulate_portfolio_orders(
    *,
    run_id: str,
    events: list[dict[str, Any]],
    rules: list[dict[str, Any]],
    max_holding_bars: int | None,
    stop_loss_qtr_multiplier: float,
    entry_condition: str = "mixed",
    initial_capital: Decimal = Decimal("1000000"),
    single_symbol_position_pct: Decimal | None = Decimal("10"),
    position_sizing_mode: str = "PERCENT",
    single_symbol_lots: int | None = None,
    no_overnight: bool = False,
) -> list[dict[str, Any]]:
    """Simulate each rule with one shared capital pool across all selected products."""
    ordered_events = sorted(
        events,
        key=lambda event: (
            _signal_timestamp(event["signal"]) or pd.Timestamp.max,
            str(event["signal"].get("timeframe", "")),
            _signal_key(event["signal"]),
        ),
    )
    orders: list[dict[str, Any]] = []
    for rule in rules:
        active_positions: dict[str, dict[str, Any]] = {}
        reserved_margin = Decimal("0")
        previous_by_head: dict[str, dict[str, Any]] = {}
        for event in ordered_events:
            signal = event["signal"]
            signal_time = _signal_timestamp(signal)
            if signal_time is None:
                continue

            for product, active in list(active_positions.items()):
                exit_time = active["exit_time"]
                if exit_time is not None and signal_time > exit_time:
                    reserved_margin -= active["margin"]
                    del active_positions[product]

            product_key = _position_key(str(signal.get("symbol", "")))
            if product_key in active_positions:
                continue

            head_key = f"{product_key}|{_head_key(signal)}"
            previous = previous_by_head.get(head_key)
            if previous is not None and previous["exit_reason"] == "STOP_LOSS" and not _has_higher_score(signal, previous["signal"]):
                continue

            order = _simulate_order(
                run_id=run_id,
                series_id=event["series_id"],
                frame=event["frame"],
                signal=signal,
                rule=rule,
                max_holding_bars=max_holding_bars,
                contract=event["contract"],
                entry_condition=entry_condition,
                stop_loss_qtr_multiplier=stop_loss_qtr_multiplier,
                initial_capital=initial_capital,
                single_symbol_position_pct=single_symbol_position_pct,
                position_sizing_mode=position_sizing_mode,
                single_symbol_lots=single_symbol_lots,
                no_overnight=no_overnight,
            )
            if order["status"] == "INVALID":
                continue

            margin = _position_margin(order, event["contract"])
            if reserved_margin + margin > initial_capital:
                continue

            previous_by_head[head_key] = {"signal": signal, "exit_reason": order["exit_reason"]}
            exit_time = pd.Timestamp(order["exit_time"]) if order["exit_time"] is not None else None
            active_positions[product_key] = {"exit_time": exit_time, "margin": margin}
            reserved_margin += margin
            orders.append(order)
    return orders


def _simulate_symbol_orders(**kwargs: Any) -> list[dict[str, Any]]:
    """Backward-compatible name for the shared-capital order simulator."""
    return _simulate_portfolio_orders(**kwargs)


def _entry_condition_streams(conditions: list[str]) -> tuple[list[str], set[str]]:
    normalized = [condition for condition in conditions if ":" in condition]
    trigger_keys = list(dict.fromkeys(
        condition.rsplit(":", 1)[-1]
        for condition in normalized
        if condition.rsplit(":", 1)[-1] in {"right_shoulder_confirmed", "right_neck_confirmed"}
    ))
    pullback_types = {
        condition.rsplit(":", 1)[-1]
        for condition in normalized
        if condition.rsplit(":", 1)[-1].endswith("_pullback")
    }
    if not trigger_keys and pullback_types:
        trigger_keys = ["pullback"]
    return trigger_keys, pullback_types


def _summaries(
    run_id: str,
    rules: list[dict[str, Any]],
    orders: list[dict[str, Any]],
    entry_conditions: list[str] | None = None,
) -> list[dict[str, Any]]:
    trigger_keys, pullback_types = _entry_condition_streams(entry_conditions or [])
    if not trigger_keys:
        trigger_keys = list(dict.fromkeys(
            str(item.get("entry_condition") if item.get("entry_condition") not in {None, "", "mixed"} else item.get("alert_type") or "")
            for item in orders
            if str(item.get("entry_condition") if item.get("entry_condition") not in {None, "", "mixed"} else item.get("alert_type") or "") in {"right_shoulder_confirmed", "right_neck_confirmed"}
        ))

    rows: list[dict[str, Any]] = []
    for rule in rules:
        for entry_condition in trigger_keys:
            grouped = [
                item for item in orders
                if item["rule_key"] == rule["key"]
                and (
                    (item.get("entry_condition") not in {None, "", "mixed"} and item.get("entry_condition") == entry_condition)
                    or (
                        item.get("entry_condition") in {None, "", "mixed"}
                        and (
                            item.get("alert_type") == entry_condition
                            or (entry_condition != "pullback" and item.get("alert_type") in pullback_types)
                            or (entry_condition == "pullback" and item.get("alert_type") in pullback_types)
                        )
                    )
                )
                and item["status"] != "INVALID"
            ]
            closed = [item for item in grouped if item["status"] == "CLOSED"]
            outcomes = [
                Decimal(str((item["net_pnl"] if item["cost_available"] else item["r_multiple"]) or 0))
                for item in closed
            ]
            wins = sum(value > 0 for value in outcomes)
            losses = sum(value < 0 for value in outcomes)
            breakevens = sum(value == 0 for value in outcomes)
            rs = [Decimal(str(item["r_multiple"] or 0)) for item in closed]
            positives = sum((value for value in rs if value > 0), Decimal("0"))
            negatives = abs(sum((value for value in rs if value < 0), Decimal("0")))
            costed = [item for item in closed if item["cost_available"]]
            rows.append({
                "run_id": run_id,
                "rule_key": rule["key"],
                "rule_label": rule["label"],
                "entry_condition": entry_condition,
                "rule_type": rule["type"],
                "multiplier": rule.get("multiplier"),
                "sample_count": len(grouped),
                "wins": wins,
                "losses": losses,
                "breakevens": breakevens,
                "incomplete": sum(item["status"] == "INCOMPLETE" for item in grouped),
                "take_profit_hits": sum(item["exit_reason"] == "TAKE_PROFIT" for item in closed),
                "stop_hits": sum(item["exit_reason"] == "STOP_LOSS" for item in closed),
                "time_exits": sum(item["exit_reason"] in {"TIME_EXIT", "SESSION_EXIT"} for item in closed),
                "win_rate": Decimal(wins) / Decimal(wins + losses) if wins + losses else Decimal("0"),
                "gross_pnl": sum((Decimal(str(item["gross_pnl"])) for item in costed), Decimal("0")) if costed else None,
                "net_pnl": sum((Decimal(str(item["net_pnl"])) for item in costed), Decimal("0")) if costed else None,
                "avg_r": sum(rs, Decimal("0")) / Decimal(len(rs)) if rs else Decimal("0"),
                "total_r": sum(rs, Decimal("0")),
                "profit_factor": positives / negatives if negatives > 0 else None,
                "avg_holding_bars": Decimal(sum(item["holding_bars"] for item in closed)) / Decimal(len(closed)) if closed else Decimal("0"),
            })
    return rows


async def process_next_backtest_run(worker_id: str) -> bool:
    run = claim_next_backtest_run(worker_id)
    if run is None:
        return False
    run_id = run["id"]
    request = run["request"]
    all_orders: list[dict[str, Any]] = []
    signal_count = 0
    completed = 0
    errors = 0
    heartbeat_stop = asyncio.Event()
    lease_lost = asyncio.Event()
    heartbeat_task = asyncio.create_task(
        _heartbeat_backtest_run(run_id, worker_id, heartbeat_stop, lease_lost),
        name=f"backtest-heartbeat-{run_id}",
    )
    try:
        all_events: list[dict[str, Any]] = []
        for symbol in request["symbols"]:
            symbol_events: list[dict[str, Any]] = []
            for timeframe in request["timeframes"]:
                if lease_lost.is_set():
                    raise BacktestLeaseLostError(f"backtest lease lost: {run_id}")
                if is_backtest_cancel_requested(run_id):
                    replace_backtest_summaries(
                        run_id,
                        _summaries(
                            run_id,
                            request["take_profit_rules"],
                            all_orders,
                            [*request.get("entry_conditions", []), *request.get("other_entry_conditions", [])],
                        ),
                    )
                    finish_backtest_run(
                        run_id, "CANCELLED", signal_count, len(all_orders), worker_id=worker_id,
                    )
                    return True
                try:
                    support_limit = max(240, min(int(request["kline_count"]), 600))
                    frame, hourly, daily = await asyncio.gather(
                        fetch_kline_from_market(symbol, timeframe, int(request["kline_count"])),
                        fetch_kline_from_market(symbol, "1h", support_limit),
                        fetch_kline_from_market(symbol, "1d", support_limit),
                    )
                    frame, selected, chart = await asyncio.to_thread(
                        _analyze_market,
                        frame,
                        hourly,
                        daily,
                        symbol,
                        timeframe,
                        request["entry_conditions"],
                        request["other_entry_conditions"],
                        int(request["min_pattern_score"]),
                        int(request["min_trend_score"]),
                        int(request["other_min_pattern_score"]),
                        int(request["other_max_trend_score"]),
                    )
                    series_id = save_backtest_series(run_id, symbol, timeframe, {
                        "symbol": symbol, "timeframe": timeframe, "chart": chart, "signals": selected,
                    })
                    contract = _contract_for_symbol(symbol)
                    symbol_events.extend({
                        "series_id": series_id,
                        "frame": frame,
                        "signal": signal,
                        "contract": contract,
                    } for signal in selected)
                    signal_count += len(selected)
                except Exception as exc:
                    errors += 1
                    logger.exception("backtest combination failed: run=%s symbol=%s timeframe=%s", run_id, symbol, timeframe)
                    add_backtest_error(run_id, symbol, timeframe, str(exc))
                completed += 1
                update_backtest_progress(
                    run_id, completed, signal_count, len(all_orders), worker_id=worker_id,
                )
            all_events.extend(symbol_events)
        configured_conditions = [*request.get("entry_conditions", []), *request.get("other_entry_conditions", [])]
        entry_streams, pullback_types = _entry_condition_streams(configured_conditions)
        for entry_condition in entry_streams:
            stream_events = [
                event for event in all_events
                if (
                    event["signal"].get("alert_type") == entry_condition
                    or (
                        entry_condition != "pullback"
                        and event["signal"].get("alert_type") in pullback_types
                    )
                    or (
                        entry_condition == "pullback"
                        and event["signal"].get("alert_type") in pullback_types
                    )
                )
            ]
            if not stream_events:
                continue
            portfolio_orders = await asyncio.to_thread(
                _simulate_portfolio_orders,
                run_id=run_id,
                events=stream_events,
                rules=request["take_profit_rules"],
                max_holding_bars=(
                    int(request["max_holding_bars"])
                    if request.get("max_holding_bars") is not None
                    else None
                ),
                stop_loss_qtr_multiplier=float(request.get("stop_loss_qtr_multiplier", 0.5)),
                entry_condition=entry_condition,
                initial_capital=Decimal(str(request.get("initial_capital", 1_000_000))),
                single_symbol_position_pct=(
                    Decimal(str(request["single_symbol_position_pct"]))
                    if request.get("single_symbol_position_pct") is not None
                    else None
                ),
                position_sizing_mode=str(request.get("position_sizing_mode", "PERCENT")),
                single_symbol_lots=(
                    int(request["single_symbol_lots"])
                    if request.get("single_symbol_lots") is not None
                    else None
                ),
                no_overnight=bool(request.get("no_overnight", False)),
            )
            save_backtest_orders(portfolio_orders)
            all_orders.extend(portfolio_orders)
            update_backtest_progress(
                run_id, completed, signal_count, len(all_orders), worker_id=worker_id,
            )
        replace_backtest_summaries(
            run_id,
            _summaries(
                run_id,
                request["take_profit_rules"],
                all_orders,
                [*request.get("entry_conditions", []), *request.get("other_entry_conditions", [])],
            ),
        )
        status = "COMPLETED_WITH_ERRORS" if errors else "COMPLETED"
        finish_backtest_run(run_id, status, signal_count, len(all_orders), worker_id=worker_id)
    except BacktestLeaseLostError:
        logger.warning("backtest worker lost task ownership: run=%s worker=%s", run_id, worker_id)
    except Exception as exc:
        logger.exception("backtest run failed: run=%s", run_id)
        finish_backtest_run(
            run_id, "FAILED", signal_count, len(all_orders), str(exc), worker_id=worker_id,
        )
    finally:
        heartbeat_stop.set()
        await heartbeat_task
    return True


def build_backtest_export(user_id: int, run_id: str) -> tuple[BytesIO, str]:
    run = get_backtest_run(user_id, run_id)
    orders = all_backtest_orders(user_id, run_id)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "回测参数"
    summary_sheet.append(["字段", "值"])
    summary_sheet.append(["名称", run["name"]])
    summary_sheet.append(["状态", run["status"]])
    for key, value in run["request"].items():
        summary_sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value])

    rule_sheet = workbook.create_sheet("止盈条件对比")
    rule_columns = ["rule_label", "entry_condition", "sample_count", "wins", "losses", "breakevens", "incomplete", "win_rate", "take_profit_hits", "stop_hits", "time_exits", "net_pnl", "avg_r", "total_r", "profit_factor", "avg_holding_bars"]
    rule_sheet.append(rule_columns)
    for row in run["summaries"]:
        rule_sheet.append([row.get(column) for column in rule_columns])

    order_sheet = workbook.create_sheet("订单详情")
    order_columns = ["symbol", "timeframe", "rule_label", "pattern", "alert_type", "direction", "quantity", "status", "exit_reason", "entry_time", "exit_time", "entry_price", "stop_price", "target_price", "exit_price", "net_pnl", "fees", "r_multiple", "holding_bars", "mfe_r", "mae_r"]
    order_sheet.append(order_columns)
    for row in orders:
        order_sheet.append([row.get(column) for column in order_columns])

    error_sheet = workbook.create_sheet("失败组合")
    error_sheet.append(["symbol", "timeframe", "message"])
    for row in run["errors"]:
        error_sheet.append([row["symbol"], row["timeframe"], row["message"]])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, f"backtest-{run_id}.xlsx"
